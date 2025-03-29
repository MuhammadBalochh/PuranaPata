import json
from fpdf import FPDF
import openpyxl

class Exporters:
    @staticmethod
    def export_to_pdf(url, snapshots):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Snapshots for {url}", ln=True)
        for timestamp, snapshot in snapshots:
            pdf.cell(200, 10, txt=f"{timestamp}: {snapshot}", ln=True)
        pdf.output("output.pdf")

    @staticmethod
    def export_to_json(url, snapshots):
        with open(f"{url}_snapshots.json", "w") as f:
            json.dump(snapshots, f)

    @staticmethod
    def export_to_excel(url, snapshots):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Snapshots of {url}"
        for i, (timestamp, snapshot) in enumerate(snapshots, start=1):
            ws.append([timestamp, snapshot])
        wb.save(f"{url}_snapshots.xlsx")
